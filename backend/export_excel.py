import pandas as pd
from sqlalchemy import create_engine
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 1. Conexión a tu base de datos SQLite
engine = create_engine('sqlite:///marina_encuestas.db')

def generar_reporte_excel():
    # 2. Cargar datos usando SQL
    # Unimos con la tabla departamentos para tener los nombres reales
    query = """
    SELECT e.id, e.fecha_creacion, d.nombre as departamento, 
           e.p1_trato, e.p2_efectividad, e.p3_facilidad_reporte, 
           e.p4_calidad_conexion, e.p5_estado_equipos, e.p6_comunicacion, 
           e.comentarios
    FROM encuestas e
    JOIN departamentos d ON e.departamento_id = d.id
    """
    df = pd.read_sql(query, engine)

    # 3. Calcular KPIs rápidos
    kpis = df.groupby('departamento').agg({
        'p1_trato': 'mean',
        'p2_efectividad': 'mean',
        'p4_calidad_conexion': 'mean',
        'id': 'count'
    }).rename(columns={'id': 'Total_Encuestas'}).reset_index()

    # 4. Exportar con formato
    filename = "Reporte_KPIs_Marina.xlsx"
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Datos_Crudos', index=False)
        kpis.to_excel(writer, sheet_name='Resumen_KPIs', index=False)
        
        # Aplicar estilos básicos al encabezado
        workbook = writer.book
        ws = workbook['Resumen_KPIs']
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    print(f"✅ Reporte generado: {filename}")

if __name__ == "__main__":
    generar_reporte_excel()